import os
import time
from datetime import datetime
import pandas as pd

# --- Selenium and WebDriver Imports ---
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# --- openpyxl for postprocessing the Excel file ---
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#########################################
# STEP 1. Download the file via Selenium
#########################################

# Set up the download directory
download_dir = os.path.join(os.getcwd(), "downloads")
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

chrome_options = Options()
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)
# Uncomment the next line to run headless if desired:
# chrome_options.headless = True

driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

# URL for TRUMBULL county file (which gives a 403 when using requests)
download_url = "https://www6.ohiosos.gov/ords/f?p=VOTERFTP:DOWNLOAD::FILE:NO:2:P2_PRODUCT_NUMBER:78"
driver.get(download_url)

timeout = 60  # seconds
start_time = time.time()
downloaded_file_path = None

while time.time() - start_time < timeout:
    files = os.listdir(download_dir)
    txt_files = [f for f in files if f.lower().endswith('.txt')]
    if txt_files:
        downloaded_file_path = os.path.join(download_dir, txt_files[0])
        break
    time.sleep(1)

if not downloaded_file_path:
    print("File download timed out. Exiting.")
    driver.quit()
    exit()

print(f"Downloaded file: {downloaded_file_path}")
driver.quit()

######################################################
# STEP 2. Process the downloaded file with pandas
######################################################
try:
    df = pd.read_csv(downloaded_file_path, delimiter=",")
except Exception as e:
    print(f"Error reading the file with pandas: {e}")
    exit()

# Normalize column names (trim and convert to uppercase)
df.columns = df.columns.str.strip().str.upper()

# Filter rows: use the WARD column to include only rows containing "WARREN-WARD"
filtered_df = df[df["WARD"].str.contains("WARREN-WARD", case=False, na=False)]

# Sort by PRECINCT_NAME
try:
    sorted_df = filtered_df.sort_values(by="PRECINCT_NAME")
except KeyError:
    print("The column 'PRECINCT_NAME' was not found. Check the file headers.")
    print("Available columns:", filtered_df.columns.tolist())
    exit()

#############################################
# Helper function: Post-process an Excel file
#############################################
def postprocess_excel(filename):
    """
    Open the Excel file, insert 6 new columns immediately to the right of the WARD column
    with the following headers and formulas:
      1. Total:  =COUNTA(vote_range)
      2. Dems:   =COUNTIF(vote_range,"D")
      3. REPS:   =COUNTIF(vote_range,"R")
      4. Muni:   =SUM(IF(...)) for odd-year vote columns (if cell="D", count 1)
      5. Latest: =SUM(IF(...)) for vote columns in the last 6 years (if cell<>"", count 1)
      6. Both:   =IF(AND(Dems_cell>0, REPS_cell>0),1,0)
    Then, insert a column after FIRST_NAME named DISPLAY that concatenates:
         LAST_NAME + " " + LEFT(DATE_OF_BIRTH,4) + "T=" + Total + "D=" + Dems + "R=" + REPS + "M=" + Muni + "L=" + Latest + "B=" + Both
    Also, insert a column after RESIDENTIAL_ADDRESS1 named StreetName that extracts the street name.
    """
    wb = load_workbook(filename)
    ws = wb.active

    # -------------------------------
    # (A) Insert 6 columns after WARD
    # -------------------------------
    ward_col_idx = None
    primary_col_idx = None  # first vote column (assumed to be labeled "PRIMARY-03/07/2000")
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if isinstance(cell_val, str):
            val = cell_val.strip().upper()
            if val == "WARD":
                ward_col_idx = col
            if val == "PRIMARY-03/07/2000":
                primary_col_idx = col

    if ward_col_idx is None:
        print("Could not find 'WARD' column in the header of", filename)
        return
    if primary_col_idx is None:
        print("Could not find 'PRIMARY-03/07/2000' column in the header of", filename)
        return

    # Insert 6 new columns immediately to the right of WARD.
    insert_position = ward_col_idx + 1
    ws.insert_cols(insert_position, amount=6)
    # Adjust primary vote column index:
    new_primary_idx = primary_col_idx + 6

    # Identify the last vote column (assumed to be the last column in the worksheet).
    last_vote_col_idx = ws.max_column

    # Determine odd-year vote columns (for "Muni")
    odd_year_cols = []
    for col in range(new_primary_idx, last_vote_col_idx + 1):
        header_val = ws.cell(row=1, column=col).value
        if isinstance(header_val, str) and header_val.upper().startswith("PRIMARY-"):
            try:
                year = int(header_val.strip()[-4:])
                if year % 2 == 1:
                    odd_year_cols.append(col)
            except:
                pass
    odd_year_letters = [get_column_letter(c) for c in odd_year_cols]

    # Determine "Latest" vote columns (for votes in the last 6 years).
    # Compute the threshold: any vote column with a header date with year >= (current_year - 6)
    current_year = datetime.today().year
    latest_threshold = current_year - 6  # e.g., if current_year=2025, include columns with year >= 2019
    latest_cols = []
    for col in range(new_primary_idx, last_vote_col_idx + 1):
        header_val = ws.cell(row=1, column=col).value
        if isinstance(header_val, str) and (("PRIMARY-" in header_val.upper()) or ("SPECIAL-" in header_val.upper())):
            try:
                year = int(header_val.strip()[-4:])
                if year >= latest_threshold:
                    latest_cols.append(col)
            except:
                pass
    latest_letters = [get_column_letter(c) for c in latest_cols]

    # Write the new header labels for the 6 inserted columns.
    new_headers = ["Total:", "Dems", "REPS", "Muni", "Latest", "Both"]
    for i, header in enumerate(new_headers):
        ws.cell(row=1, column=insert_position + i, value=header)

    # For each data row, add formulas in the 6 inserted columns.
    first_vote_letter = get_column_letter(new_primary_idx)
    last_vote_letter = get_column_letter(last_vote_col_idx)
    for row in range(2, ws.max_row + 1):
        vote_range = f"${first_vote_letter}${row}:${last_vote_letter}${row}"
        # Total: count nonblank cells in the vote range.
        total_formula = f"=COUNTA({vote_range})"
        ws.cell(row=row, column=insert_position, value=total_formula)

        # Dems: count cells equal to "D".
        dems_formula = f'=COUNTIF({vote_range},"D")'
        ws.cell(row=row, column=insert_position + 1, value=dems_formula)

        # REPS: count cells equal to "R".
        reps_formula = f'=COUNTIF({vote_range},"R")'
        ws.cell(row=row, column=insert_position + 2, value=reps_formula)

        # Muni: sum IF statements for odd-year vote columns.
        muni_parts = []
        for col_letter in odd_year_letters:
            muni_parts.append(f'IF({col_letter}{row}="D",1,0)')
        muni_formula = "=" + "+".join(muni_parts) if muni_parts else "=0"
        ws.cell(row=row, column=insert_position + 3, value=muni_formula)

        # Latest: sum IF statements for vote columns with header year >= latest_threshold.
        latest_parts = []
        for col_letter in latest_letters:
            latest_parts.append(f'IF({col_letter}{row}<>"",1,0)')
        latest_formula = "=" + "+".join(latest_parts) if latest_parts else "=0"
        ws.cell(row=row, column=insert_position + 4, value=latest_formula)

        # Both: if Dems > 0 AND REPS > 0 then 1, else 0.
        dem_cell = get_column_letter(insert_position + 1) + str(row)
        rep_cell = get_column_letter(insert_position + 2) + str(row)
        both_formula = f"=IF(AND({dem_cell}>0, {rep_cell}>0),1,0)"
        ws.cell(row=row, column=insert_position + 5, value=both_formula)

    # -------------------------------
    # (B) Insert a "DISPLAY" column after FIRST_NAME
    # -------------------------------
    first_name_idx = None
    last_name_idx = None
    dob_idx = None
    # Also locate RESIDENTIAL_ADDRESS1 for later use.
    res_address_idx = None
    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if cell_val:
            header = str(cell_val).strip().upper()
            if header == "FIRST_NAME":
                first_name_idx = col
            elif header == "LAST_NAME":
                last_name_idx = col
            elif header == "DATE_OF_BIRTH":
                dob_idx = col
            elif header == "RESIDENTIAL_ADDRESS1":
                res_address_idx = col

    if first_name_idx is not None:
        ws.insert_cols(first_name_idx + 1)
        ws.cell(row=1, column=first_name_idx + 1, value="DISPLAY")
        for row in range(2, ws.max_row + 1):
            # Build cell references for LAST_NAME and DATE_OF_BIRTH.
            last_name_cell = get_column_letter(last_name_idx) + str(row) if last_name_idx else ""
            dob_cell = get_column_letter(dob_idx) + str(row) if dob_idx else ""
            # The inserted vote columns (Total, Dems, REPS, Muni, Latest, Both) are located after WARD.
            # Their positions relative to WARD: Total at (ward_col_idx+1), Dems at (ward_col_idx+2),
            # REPS at (ward_col_idx+3), Muni at (ward_col_idx+4), Latest at (ward_col_idx+5), Both at (ward_col_idx+6).
            total_cell = get_column_letter(ward_col_idx + 1) + str(row)
            dems_cell = get_column_letter(ward_col_idx + 2) + str(row)
            reps_cell = get_column_letter(ward_col_idx + 3) + str(row)
            muni_cell = get_column_letter(ward_col_idx + 4) + str(row)
            latest_cell = get_column_letter(ward_col_idx + 5) + str(row)
            both_cell = get_column_letter(ward_col_idx + 6) + str(row)
            display_formula = (
                f'=CONCATENATE({last_name_cell}," ",LEFT({dob_cell},4),"T=",'
                f'{total_cell},"D=",{dems_cell},"R=",{reps_cell},"M=",{muni_cell},"L=",{latest_cell},"B=",{both_cell})'
            )
            ws.cell(row=row, column=first_name_idx + 1, value=display_formula)

    # -------------------------------
    # (C) Insert a "StreetName" column after RESIDENTIAL_ADDRESS1
    # -------------------------------
    if res_address_idx is not None:
        ws.insert_cols(res_address_idx + 1)
        ws.cell(row=1, column=res_address_idx + 1, value="StreetName")
        for row in range(2, ws.max_row + 1):
            address_cell = get_column_letter(res_address_idx) + str(row)
            street_formula = f'=RIGHT({address_cell},LEN({address_cell})-FIND(" ",{address_cell}))'
            ws.cell(row=row, column=res_address_idx + 1, value=street_formula)
            
    wb.save(filename)
    print(f"Post-processing complete. Final file saved as {filename}")

#############################################
# STEP 3. Write the Overall DataFrame to Excel
#############################################
today_str = datetime.today().strftime("%Y-%m-%d")
overall_filename = f"CityOfWarren{today_str}.xlsx"

try:
    sorted_df.to_excel(overall_filename, index=False, engine='openpyxl')
    print(f"Overall data written to {overall_filename}")
except Exception as e:
    print(f"Error saving the overall Excel file: {e}")
    exit()

postprocess_excel(overall_filename)

#######################################################
# STEP 4. Create separate files for each unique WARD
#######################################################
unique_wards = sorted_df["WARD"].unique()

for ward in unique_wards:
    ward_df = sorted_df[sorted_df["WARD"] == ward]
    ward_filename = f"City_of_{ward}-{today_str}.xlsx"
    try:
        ward_df.to_excel(ward_filename, index=False, engine='openpyxl')
        print(f"Data for ward '{ward}' written to {ward_filename}")
    except Exception as e:
        print(f"Error saving Excel file for ward '{ward}': {e}")
        continue
    postprocess_excel(ward_filename)
