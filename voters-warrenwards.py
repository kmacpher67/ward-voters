import os
import time
from datetime import datetime
import pandas as pd

# --- Selenium and WebDriver Imports ---
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# --- openpyxl for postprocessing the Excel file ---
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#########################################
# STEP 1. Download the file via Selenium
#########################################

# Set up the download directory
download_dir = os.path.join(os.getcwd(), "downloads")
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

# Configure Chrome options to automatically download files.
chrome_options = Options()
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)
# Uncomment next line to run in headless mode if needed:
# chrome_options.headless = True

# Initialize the Selenium Chrome WebDriver.
driver = webdriver.Chrome(
    service=Service(ChromeDriverManager().install()),
    options=chrome_options
)

# URL for TRUMBULL county file (which gives a 403 using requests)
download_url = "https://www6.ohiosos.gov/ords/f?p=VOTERFTP:DOWNLOAD::FILE:NO:2:P2_PRODUCT_NUMBER:78"
driver.get(download_url)

# Wait for the file download (poll the download_dir for a .txt file)
timeout = 60  # seconds
start_time = time.time()
downloaded_file_path = None

while time.time() - start_time < timeout:
    files = os.listdir(download_dir)
    txt_files = [f for f in files if f.lower().endswith('.txt')]
    if txt_files:
        downloaded_file_path = os.path.join(download_dir, txt_files[0])
        break
    time.sleep(1)

if not downloaded_file_path:
    print("File download timed out. Exiting.")
    driver.quit()
    exit()

print(f"Downloaded file: {downloaded_file_path}")
driver.quit()

######################################################
# STEP 2. Process the downloaded file with pandas
######################################################
try:
    # Adjust delimiter if necessary (here we assume CSV format)
    df = pd.read_csv(downloaded_file_path, delimiter=",")
except Exception as e:
    print(f"Error reading the file with pandas: {e}")
    exit()

# Normalize column names: remove extra whitespace and convert to uppercase.
df.columns = df.columns.str.strip().str.upper()

# Instead of filtering by CITY, filter rows where the "WARD" column contains "WARREN-WARD"
filtered_df = df[df["WARD"].str.contains("WARREN-WARD", case=False, na=False)]

# Sort the data by "PRECINCT_NAME"
try:
    sorted_df = filtered_df.sort_values(by="PRECINCT_NAME")
except KeyError:
    print("The column 'PRECINCT_NAME' was not found. Check the file headers.")
    print("Available columns:", filtered_df.columns.tolist())
    exit()

#############################################
# Helper function: Post-process an Excel file
#############################################
def postprocess_excel(filename):
    """
    Open the Excel file (filename) with openpyxl, insert four columns immediately
    to the right of the "WARD" column, and add formulas in each row:
      - Total: counts nonblank cells in the vote columns (from the first vote column to the last)
      - Dems: counts cells equal to "D"
      - REPS: counts cells equal to "R"
      - Muni: sums IF statements for odd-year vote columns that equal "D"
    The vote columns are assumed to start at header "PRIMARY-03/07/2000".
    """
    wb = load_workbook(filename)
    ws = wb.active

    # Identify key columns based on header row (row 1)
    ward_col_idx = None
    primary_col_idx = None

    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(row=1, column=col).value
        if isinstance(cell_val, str):
            cell_val = cell_val.strip().upper()
            if cell_val == "WARD":
                ward_col_idx = col
            if cell_val == "PRIMARY-03/07/2000":
                primary_col_idx = col

    if ward_col_idx is None:
        print("Could not find 'WARD' column in the header of", filename)
        return
    if primary_col_idx is None:
        print("Could not find 'PRIMARY-03/07/2000' column in the header of", filename)
        return

    # Insert four new columns immediately to the right of the WARD column.
    insert_position = ward_col_idx + 1
    ws.insert_cols(insert_position, amount=4)

    # Because of the insertion, vote columns shift right.
    new_primary_idx = primary_col_idx + 4

    # Identify the last vote column (assumed to be the last column in the worksheet).
    last_vote_col_idx = ws.max_column

    # Determine which vote columns (from new_primary_idx to last_vote_col_idx) have an odd year.
    odd_year_cols = []
    for col in range(new_primary_idx, last_vote_col_idx + 1):
        header_val = ws.cell(row=1, column=col).value
        if isinstance(header_val, str) and header_val.upper().startswith("PRIMARY-"):
            try:
                year = int(header_val.strip()[-4:])
                if year % 2 == 1:
                    odd_year_cols.append(col)
            except:
                pass

    # Get Excel column letters for the vote range.
    first_vote_letter = get_column_letter(new_primary_idx)
    last_vote_letter = get_column_letter(last_vote_col_idx)
    odd_year_letters = [get_column_letter(c) for c in odd_year_cols]

    # Write new header labels in the inserted columns.
    new_headers = ["Total:", "Dems", "REPS", "Muni"]
    for i, header in enumerate(new_headers):
        ws.cell(row=1, column=insert_position + i, value=header)

    # For each data row, insert formulas in the new columns.
    for row in range(2, ws.max_row + 1):
        vote_range = f"${first_vote_letter}${row}:${last_vote_letter}${row}"
        # Total: count nonblank cells in the vote range.
        total_formula = f"=COUNTA({vote_range})"
        ws.cell(row=row, column=insert_position, value=total_formula)
        # Dems: count cells equal to "D".
        dems_formula = f'=COUNTIF({vote_range},"D")'
        ws.cell(row=row, column=insert_position + 1, value=dems_formula)
        # REPS: count cells equal to "R".
        reps_formula = f'=COUNTIF({vote_range},"R")'
        ws.cell(row=row, column=insert_position + 2, value=reps_formula)
        # Muni: sum IF statements for odd-year vote columns (if cell equals "D", add 1).
        muni_parts = []
        for col_letter in odd_year_letters:
            muni_parts.append(f'IF({col_letter}{row}="D",1,0)')
        muni_formula = "=" + "+".join(muni_parts) if muni_parts else "=0"
        ws.cell(row=row, column=insert_position + 3, value=muni_formula)

    wb.save(filename)
    print(f"Post-processing complete. Final file saved as {filename}")

#############################################
# STEP 3. Write the Overall DataFrame to Excel
#############################################
today_str = datetime.today().strftime("%Y-%m-%d")
overall_filename = f"CityOfWarren{today_str}.xlsx"

try:
    # Write the overall sorted DataFrame to Excel.
    sorted_df.to_excel(overall_filename, index=False, engine='openpyxl')
    print(f"Overall data written to {overall_filename}")
except Exception as e:
    print(f"Error saving the overall Excel file: {e}")
    exit()

# Post-process the overall file.
postprocess_excel(overall_filename)

#######################################################
# STEP 4. Create separate files for each unique WARD
#######################################################
unique_wards = sorted_df["WARD"].unique()

for ward in unique_wards:
    ward_df = sorted_df[sorted_df["WARD"] == ward]
    # Build a filename. (Ensure ward name is safe for filenames if necessary.)
    ward_filename = f"City of {ward}-{today_str}.xlsx"
    try:
        ward_df.to_excel(ward_filename, index=False, engine='openpyxl')
        print(f"Data for ward '{ward}' written to {ward_filename}")
    except Exception as e:
        print(f"Error saving Excel file for ward '{ward}': {e}")
        continue

    # Post-process the ward file (insert columns and formulas).
    postprocess_excel(ward_filename)
